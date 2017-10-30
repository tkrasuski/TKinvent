# -*- coding: utf-8 -*-
# this is xlsx handling module
import settings as stt
import json
from openpyxl import load_workbook

class XlHandler(object):
    def __init__(self):
        self.file = None
        self.js_data = None
        self.wb = None
        self.ws = None
        self.data =[]
    def loadFile(self, file_):
        'Loads workbook file and sets it as class '
        bfile = open(file_, 'rb')
        self.wb = load_workbook(filename=bfile,read_only=True)
    def parseFile(self):
        self.ws = self.wb.active
        for row in self.ws.iter_rows(min_row=2, max_col=5, max_row=50):
            drow = {}
            for cell in row:
                #print cell.column, ':', cell.value
                if cell.column == stt.part_no:
                    drow['part_no']=cell.value
                if cell.column == stt.description:
                    drow['description']=cell.value
                if cell.column == stt.qty:
                    drow['qty']=cell.value
            self.data.append(dict(line=drow))

    def getContent(self):
        content = dict(content=self.data)
        return content
